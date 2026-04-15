import time
import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl 
from openpyxl.styles import Alignment, Font

# --- AYARLAR (SENİN NİŞ FİLTRELERİN) ---
SAYFA_BASINA_ILAN = 48  
BEKLEME_SURESI = 8      

# URL'deki parametreler:
KATEGORI_ID = "f1379fed-67c9-4f85-aa7d-c3978dfb4f24" # Senin belirlediğin kategori
MAX_FIYAT = 150000                                   # 150.000 TL Tavan Fiyat

def main():
    print(f"🚀 TASİŞ 'NİŞ AVCI' BOTU (V8) BAŞLATILIYOR...")
    print(f"🎯 Hedef: Kategori ID '{KATEGORI_ID}' | Maks Fiyat: {MAX_FIYAT} TL")
    
    chrome_options = Options()
    chrome_options.add_argument("--disable-blink-features=AutomationControlled") 
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.add_argument("--window-size=1280,800")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    driver = webdriver.Chrome(options=chrome_options)
    driver.implicitly_wait(10)
    
    tum_veriler = []
    global_taranan_linkler = set() 
    sayfa_no = 1
    
    while True:
        # --- URL OLUŞTURMA (SENİN FİLTRELERİNLE) ---
        kac_tane_atlanacak = (sayfa_no - 1) * SAYFA_BASINA_ILAN
        
        # Senin verdiğin link yapısını dinamik hale getirdik:
        url = (f"https://eihale.gov.tr/?"
               f"skipCount={kac_tane_atlanacak}&"
               f"maxResultCount={SAYFA_BASINA_ILAN}&"
               f"categoryId={KATEGORI_ID}&"
               f"maxPrice={MAX_FIYAT}")
        
        print(f"\n🌍 SAYFA {sayfa_no} GİDİLİYOR (Skip: {kac_tane_atlanacak})...")
        driver.get(url)
        time.sleep(BEKLEME_SURESI) 

        try:
            ilan_elementleri = driver.find_elements(By.CSS_SELECTOR, "a[href*='/ihaleler/detay/']")
            yeni_sayfa_linkleri = []
            
            for elem in ilan_elementleri:
                link = elem.get_attribute("href")
                if link and link not in global_taranan_linkler:
                    yeni_sayfa_linkleri.append(link)
                    global_taranan_linkler.add(link)
            
            # İstatistik
            toplam = len(ilan_elementleri)
            yeni = len(yeni_sayfa_linkleri)
            
            print(f"   📊 Bu sayfada {toplam} ilan var. ({yeni} tanesi yeni)")

            # --- ÇIKIŞ KONTROLÜ ---
            # Filtreli aramalarda sayfa sayısı az olur. İlan bitince duralım.
            if len(ilan_elementleri) == 0:
                print("🛑 ARADIĞINIZ KRİTERLERDE BAŞKA İLAN YOK. BİTTİ.")
                break 

        except Exception as e:
            print(f"   ❌ Hata: {e}")
            break

        # --- Sadece YENİ İlanları Gez ---
        for sira, link in enumerate(yeni_sayfa_linkleri, 1):
            print(f"   👉 [Sayfa {sayfa_no} | {sira}/{yeni}] Veri çekiliyor...")
            try:
                driver.get(link)
                time.sleep(3) 

                try: baslik = driver.find_element(By.CSS_SELECTOR, "h2.uppercase").text
                except: baslik = "Başlık Yok"

                try:
                    items = driver.find_elements(By.CLASS_NAME, "item")
                    fiyat = "0"
                    for item in items:
                        if "Başlangıç Bedeli" in item.text:
                            fiyat = item.text.replace("Başlangıç Bedeli:", "").strip()
                            break
                except: fiyat = "0"

                try:
                    detay = driver.find_element(By.CSS_SELECTOR, ".mat-expansion-panel-body").text.strip()
                except:
                    detay = driver.find_element(By.TAG_NAME, "body").text[:500]

                print(f"      ✅ {baslik[:30]}... ({fiyat})")

                tum_veriler.append({
                    "Sayfa": sayfa_no,
                    "Link": link,
                    "Başlık": baslik,
                    "Fiyat": fiyat,
                    "Detaylar": detay
                })

            except Exception as e:
                print(f"      ⚠️ İlan hatası: {e}")
                continue
        
        # Yedekle
        if tum_veriler:
            pd.DataFrame(tum_veriler).to_excel("tasis_nis_yedek.xlsx", index=False)
        
        sayfa_no += 1

    driver.quit()
    print("\n✅ TARAMA TAMAMLANDI. RAPOR FORMATLANIYOR...")

    # --- FORMATLAMA ---
    if tum_veriler:
        dosya_adi = "tasis_nis_firsatlar.xlsx"
        df = pd.DataFrame(tum_veriler)
        df.to_excel(dosya_adi, index=False)

        wb = openpyxl.load_workbook(dosya_adi)
        ws = wb.active
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 100

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter == 'E': 
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                else:
                    cell.alignment = Alignment(vertical='top', horizontal='left')

        wb.save(dosya_adi)
        print(f"🏆 DOSYA HAZIR: {dosya_adi}")
    else:
        print("❌ Bu kriterlere uygun hiç ilan bulunamadı.")
    
    input("Çıkmak için Enter'a basın...")

if __name__ == "__main__":
    main()